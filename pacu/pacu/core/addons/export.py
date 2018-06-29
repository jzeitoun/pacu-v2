import numpy as np
import scipy.io as scio
import zipfile
import json
import os

from cStringIO import StringIO
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font

from pacu.core.io.scanbox.method.sfreq.fit_tf import tf_stats
from pacu.core.io.scanbox.method.orientation.bestpref_tf import tf_bestpref
from . import extract_traces as et

class Export(object):
    def __init__(self, io, ws, condition, ids): #, rois):
        self.io = io
        self.ws = ws
        self.condition = condition
        self.flicker = any([trial['flicker'] for trial in self.condition.trial_list])
        self.blank = any([trial['blank'] for trial in self.condition.trial_list])
        self.ids = map(int, ids.split(','))
        self.rois = self.io.condition.workspaces.filter_by(name=self.ws)[0].rois
        self.roi_dict = {'{}{}'.format('cell_id_', roi.id) : roi.serialize() for roi in self.rois}

    def blank_responses(self, roi):
        if self.blank:
            blank_responses = [trial.attributes['value']['on']
                                for trial in roi.dttrialdff0s
                                if trial.attributes['trial_blank'] == True]
            return np.mean(blank_responses, 1)
        else:
            return []

    def flicker_responses(self, roi):
        if self.flicker:
            flicker_responses = [trial.attributes['value']['on']
                                for trial in roi.dttrialdff0s
                                if trial.attributes['trial_flicker'] == True]
            return np.mean(flicker_responses, 1)
        else:
            return []

    def traces_json(self):
        channel_map = {
                1: {'green': self.io.ch0, 'red': self.io.ch1},
                2: {'green': self.io.ch0},
                3: {'red': self.io.ch0}
                }
        channel_layout = channel_map[self.io.mat.channels]

        filtered_rois = [roi for roi in self.rois if roi.id in self.ids]
        roi_map = {roi.id: roi for roi in filtered_rois}
        roi_output = []

        for roi in filtered_rois:
            roi_output.append({
                'roi_id': roi.id,
                'polygon': et.array_to_polygon(roi.polygon)
                })

        for channel in channel_layout.keys():
            data = channel_layout[channel].mmap
            for roi in roi_output:
                point_pairs = et.polygon_to_array(roi['polygon'])
                if self.io.mat.channels == 1 and channel == 'red':
                    masked_trace = et.extract_mean_trace(data, point_pairs)
                    trace = masked_trace.data.tolist()
                else:
                    trace = roi_map[roi['roi_id']].dtoverallmean.value
                roi.update({channel: trace})

        sio = StringIO()
        json.dump(roi_output, sio)
        return sio.getvalue()

    def both(self):
        # Create in-memory zipfile
        zstring = StringIO()
        zfile = zipfile.ZipFile(zstring, 'w', allowZip64 = True)
        matlab_file = self.matlab()
        excel_file = self.excel()
        # Write files to in-memory zipfile
        zfile.writestr('data.xlsx', excel_file)
        zfile.writestr('data.mat', matlab_file)
        zfile.close()
        return zstring.getvalue()

    def matlab(self):
        merged_dict = {}
        # providing filename and workspace name as list to maintain compatibility
        # with post analysis scripts
        ########################################
        merged_dict['filenames'] = [str(self.io.path)]
        merged_dict['workspaces'] = [str(self.ws)]
        ########################################
        merged_dict['rois'] = self.roi_dict
        filtered_rois = [roi for roi in self.rois if roi.id in self.ids]
        for roi in filtered_rois:
            #merged_dict['rois']['{}{}'.format('cell_id_', roi.params.cell_id)]['sorted_dtorientationsmeans'] = self.sorted_orientation_traces(roi)
            merged_dict['rois']['{}{}'.format('cell_id_', roi.id)]['blank_responses'] = self.blank_responses(roi)
            merged_dict['rois']['{}{}'.format('cell_id_', roi.id)]['flicker_responses'] = self.flicker_responses(roi)
        sio = StringIO()
        scio.savemat(sio, {'merged_dict': merged_dict})
        return sio.getvalue()
        #if filename == None:
        #    fname = self.fw_array[0][0][:-3] + '_merged.mat'
        #else:
        #    fname = filename + '.mat'
        ## save .mat file
        #sio.savemat(fname,{'merged_dict':merged_dict})


    def excel(self, p_value=0.01, unmerge=0):
        '''
        Export data in Excel format (.xlsx)
        '''
        filtered_rois = [roi for roi in self.rois if roi.id in self.ids]
        num_rois = len(filtered_rois)
        sfreqs = self.condition.sfrequencies
        tfreqs = self.condition.tfrequencies
        num_sf = len(sfreqs)
        idx_list = range(3, num_rois*num_sf, num_sf)
        font = 'Courier New'

        # Create in-memory zipfile
        #zstring = StringIO()
        #zfile = zipfile.ZipFile(zstring, 'w', compression=zipfile.ZIP_DEFLATED)
        sio = StringIO()

        # Create new workbook
        wb = Workbook()
        wb.remove(wb.worksheets[0]) # remove first blank sheet

        # Create cell formats
        header = NamedStyle(name='header')
        header.alignment = Alignment(horizontal='center',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=True,
                                        shrink_to_fit=False,
                                        indent=0)

        header.font = Font(name=font,
                                size=10,
                                bold=True,
                                italic=False,
                                vertAlign=None,
                                underline='none',
                                strike=False,
                                color='FF000000')

        header.border = Border(top=Side(border_style='medium',
                                        color='FF000000'),
                                    bottom=Side(border_style='medium',
                                        color='FF000000')
                                    )

        reg_cell = NamedStyle(name='regular')
        reg_cell.alignment = Alignment(horizontal=None,
                                     vertical='center',
                                     text_rotation=0,
                                     wrap_text=True,
                                     shrink_to_fit=False,
                                     indent=0)

        reg_cell.font = Font(name=font,size=10)

        reg_cell.border = Border(top=Side(border_style='medium',
                                                 color='FF000000'),
                                             bottom=Side(border_style='medium',
                                                 color='FF000000')
                                             )

        sig_cell = NamedStyle(name='significant')
        sig_cell.alignment = Alignment(horizontal=None,
                                          vertical='center',
                                          text_rotation=0,
                                          wrap_text=True,
                                          shrink_to_fit=False,
                                          indent=0)

        sig_cell.font = Font(name=font,size=10)

        sig_cell.border = Border(top=Side(border_style='medium',
                                                 color='FF000000'),
                                             bottom=Side(border_style='medium',
                                                 color='FF000000')
                                             )

        sig_cell.fill = PatternFill(start_color='FFFFFF00',
                                        end_color='FFFFFF00',
                                        fill_type='solid')

        # Create new worksheet for each temporal frequency
        for tfreq in tfreqs:
            ws = wb.create_sheet('TFreq {}'.format(tfreq))

            # format header columns
            ws.merge_cells('A1:A2')
            ws.merge_cells('B1:C1')
            ws.merge_cells('D1:E1')
            ws.merge_cells('F1:I1')
            ws.merge_cells('S1:T1')

            for top,bottom in zip(ws['J1:Q1'][0],ws['J2:Q2'][0]):
                ws.merge_cells('{}{}:{}{}'.format(top.column,top.row,bottom.column,bottom.row))

            # write column titles
            ws['A1'].value = 'Cell ID'
            ws['A1'].style = header
            ws['B1'].value = 'Anova All'
            ws['B1'].style = header
            ws['D1'].value = 'SF Cutoff Rel33'
            ws['D1'].style = header
            ws['F1'].value = 'SF'
            ws['F1'].style = header
            ws['S1'].value = 'Anova Each'
            ws['S1'].style = header
            ws['S2'].value = 'F'
            ws['S2'].style = header
            ws['T2'].value = 'P'
            ws['T2'].style = header

            for cell,val in zip(ws[2][1:9], ['F','P','X','Y','Peak','Pref','Bandwidth','Global\nOPref']):
                cell.value = val
                cell.style = header

            for cell,val in zip(ws[1][9:18], ['@', 'OSI', 'CV', 'DCV', 'DSI', 'Sigma', 'OPref', 'RMax', 'Residual']):
                cell.value = val
                cell.style = header

            for idx,roi in zip(idx_list, filtered_rois):
                peak_sf = round(roi.dtsfreqfits.filter_by(trial_tf=tfreq).first.attributes['value']['peak'],2)
                try:
                    if roi.dtanovaeachs.filter_by(trial_tf=tfreq).filter_by(trial_sf=peak_sf).first.p <= p_value:
                        style = sig_cell
                    else:
                        style = reg_cell
                except IndexError:
                    style = reg_cell

                if unmerge == 0:
                    for top,bottom in zip(ws['A{}:I{}'.format(idx, idx)][0], ws['A{}:I{}'.format(idx+num_sf-1, idx+num_sf-1)][0]):
                        ws.merge_cells('{}{}:{}{}'.format(top.column, top.row, bottom.column, bottom.row))

                    ws.cell(row=idx,column=1).value = int(roi.id)
                    ws.cell(row=idx,column=1).style = style
                    ws.cell(row=idx,column=2).value = roi.dtanovaalls.filter_by(trial_tf=tfreq).first.attributes['value']['f']
                    ws.cell(row=idx,column=2).style = style
                    ws.cell(row=idx,column=3).value = roi.dtanovaalls.filter_by(trial_tf=tfreq).first.attributes['value']['p']
                    ws.cell(row=idx,column=3).style = style
                    try:
                        ws.cell(row=idx,column=4).value = roi.dtsfreqfits.filter_by(trial_tf=tfreq).first.attributes['value']['rc33'].x
                        ws.cell(row=idx,column=5).value = roi.dtsfreqfits.filter_by(trial_tf=tfreq).first.attributes['value']['rc33'].y
                    except AttributeError:
                        #print 'No "SF Cutoff Rel33" found for cell ',roi.params.cell_id
                        ws.cell(row=idx,column=4).value = None
                        ws.cell(row=idx,column=5).value = None
                    ws.cell(row=idx,column=4).style = style
                    ws.cell(row=idx,column=5).style = style
                    ws.cell(row=idx,column=6).value = peak_sf
                    ws.cell(row=idx,column=6).style = style
                    ws.cell(row=idx,column=7).value = roi.dtsfreqfits.filter_by(trial_tf=tfreq).first.attributes['value']['pref']
                    ws.cell(row=idx,column=7).style = style
                    ws.cell(row=idx,column=8).value = roi.dtsfreqfits.filter_by(trial_tf=tfreq).first.attributes['value']['ratio']
                    ws.cell(row=idx,column=8).style = style
                    ws.cell(row=idx,column=9).value = roi.dtorientationbestprefs.filter_by(trial_tf=tfreq).first.attributes['value']
                    ws.cell(row=idx,column=9).style = style

                elif unmerge == 1:
                    for i in range(num_sf):
                        ws.cell(row=idx+i,column=1).value = int(roi.id)
                        ws.cell(row=idx+i,column=1).style = style
                        ws.cell(row=idx+i,column=2).value = roi.dtanovaalls.filter_by(trial_tf=tfreq).first.attributes['value']['f']
                        ws.cell(row=idx+i,column=2).style = style
                        ws.cell(row=idx+i,column=3).value = roi.dtanovaalls.filter_by(trial_tf=tfreq).first.attributes['value']['p']
                        ws.cell(row=idx+i,column=3).style = style
                        try:
                            ws.cell(row=idx+i,column=4).value = roi.dtsfreqfits.filter_by(trial_tf=tfreq).first.attributes['value']['rc33'].x
                            ws.cell(row=idx+i,column=5).value = roi.dtsfreqfits.filter_by(trial_tf=tfreq).first.attributes['value']['rc33'].y
                        except AttributeError:
                            #print 'No "SF Cutoff Rel33" found for cell ',roi.params.cell_id
                            ws.cell(row=idx+i,column=4).value = None
                            ws.cell(row=idx+i,column=5).value = None
                        ws.cell(row=idx+i,column=4).style = style
                        ws.cell(row=idx+i,column=5).style = style
                        ws.cell(row=idx+i,column=6).value = peak_sf
                        ws.cell(row=idx+i,column=6).style = style
                        ws.cell(row=idx+i,column=7).value = roi.dtsfreqfits.filter_by(trial_tf=tfreq).first.attributes['value']['pref']
                        ws.cell(row=idx+i,column=7).style = style
                        ws.cell(row=idx+i,column=8).value = roi.dtsfreqfits.filter_by(trial_tf=tfreq).first.attributes['value']['ratio']
                        ws.cell(row=idx+i,column=8).style = style
                        ws.cell(row=idx+i,column=9).value = roi.dtorientationbestprefs.filter_by(trial_tf=tfreq).first.attributes['value']
                        ws.cell(row=idx+i,column=9).style = style

                for i,cell in enumerate(ws.iter_rows(min_col=10, max_col=10, min_row=idx, max_row=idx+num_sf-1)):
                        cell[0].value = sfreqs[i]
                        cell[0].style = style

                for i,row in enumerate(ws.iter_rows(min_col=11, max_col=20, min_row=idx, max_row=idx+num_sf-1)):

                    row[0].value = roi.dtorientationsfits.filter_by(trial_tf=tfreq)[i].attributes['value']['osi']
                    row[0].style = style
                    row[1].value = roi.dtorientationsfits.filter_by(trial_tf=tfreq)[i].attributes['value']['cv']
                    row[1].style = style
                    row[2].value = roi.dtorientationsfits.filter_by(trial_tf=tfreq)[i].attributes['value']['dcv']
                    row[2].style = style
                    row[3].value = roi.dtorientationsfits.filter_by(trial_tf=tfreq)[i].attributes['value']['dsi']
                    row[3].style = style
                    row[4].value = roi.dtorientationsfits.filter_by(trial_tf=tfreq)[i].attributes['value']['sigma']
                    row[4].style = style
                    row[5].value = roi.dtorientationsfits.filter_by(trial_tf=tfreq)[i].attributes['value']['o_pref']
                    row[5].style = style
                    row[6].value = roi.dtorientationsfits.filter_by(trial_tf=tfreq)[i].attributes['value']['r_max']
                    row[6].style = style
                    row[7].value = roi.dtorientationsfits.filter_by(trial_tf=tfreq)[i].attributes['value']['residual']
                    row[7].style = style
                    row[8].value = roi.dtanovaeachs.filter_by(trial_tf=tfreq)[i].attributes['f']
                    row[8].style = style
                    row[9].value = roi.dtanovaeachs.filter_by(trial_tf=tfreq)[i].attributes['p']
                    row[9].style = style

        if len(tfreqs) > 1:
            for sfreq in sfreqs:
                ws = wb.create_sheet('SFreq {}'.format(sfreq))

                # format header columns
                ws.merge_cells('A1:A2')
                ws.merge_cells('B1:C1')
                ws.merge_cells('D1:G1')

                # write column titles
                ws['A1'].value = 'Cell ID'
                ws['A1'].style = header
                ws['B1'].value = 'TF Cutoff Rel33'
                ws['B1'].style = header
                ws['D1'].value = 'TF'
                ws['D1'].style = header
                ws['G1'] # need to call to activate columns C - G

                for cell,val in zip(ws[2][1:9], ['X','Y','Peak','Pref','Bandwidth','Global\nOPref']):
                    cell.value = val
                    cell.style = header

                for idx, roi in enumerate(filtered_rois, 3):
                    style = reg_cell
                    workspace = roi.workspace
                    condition = roi.workspace.condition
                    contrast = roi.workspace.condition.contrasts[0]
                    tf_result = tf_stats(workspace, condition, roi, contrast, sfreq, dff0s=None, fits=None)
                    global_opref, _, _ = tf_bestpref(workspace, condition, roi, contrast, sfreq, dff0s=None)

                    ws.cell(row=idx,column=1).value = int(roi.id)
                    ws.cell(row=idx,column=1).style = style
                    try:
                        ws.cell(row=idx,column=2).value = tf_result['rc_33'].x
                        ws.cell(row=idx,column=3).value = tf_result['rc_33'].y
                    except:
                        #print 'No "SF Cutoff Rel33" found for cell ',roi.params.cell_id
                        ws.cell(row=idx,column=2).value = None
                        ws.cell(row=idx,column=3).value = None
                    ws.cell(row=idx,column=2).style = style
                    ws.cell(row=idx,column=3).style = style
                    ws.cell(row=idx,column=4).value = tf_result['peak']
                    ws.cell(row=idx,column=4).style = style
                    ws.cell(row=idx,column=5).value = tf_result['pref']
                    ws.cell(row=idx,column=5).style = style
                    ws.cell(row=idx,column=6).value = tf_result['ratio']
                    ws.cell(row=idx,column=6).style = style
                    ws.cell(row=idx,column=7).value = global_opref
                    ws.cell(row=idx,column=7).style = style

        # Save and return in-memory file (bytes)
            #sio = StringIO()
        wb.save(sio)
            # Write excel file to in-memory zipfile
            #zfile.writestr('tfreq_{}.xlsx'.format(tfreq), sio.getvalue())
        return sio.getvalue()

        #return zstring.getvalue()
